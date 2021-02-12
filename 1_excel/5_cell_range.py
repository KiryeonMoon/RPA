from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active
ws.title = "삼성가자"

# 한 줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])  # list나 튜플 형태
for i in range(1, 11):  # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

# col_B = ws["B"]  # B column만 가져오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

# col_range = ws["B:C"]  # "영어", "수학" 함께 가져오기
# # 위와 달리 column이 하나가 아니라 두 개 >> 반복문을 두 개 써야 함
# for cols in col_range:  # 두 개 column 중 한 개 의 column씩 가져오는중. 여기선 cols가 위의 col_B 역할을 하는 중
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1]  # 1번째 row만 가져오기
# for cell in row_title:
#     print(cell.value)

# row_range = ws[2:6]  # 1번째 줄인 title을 제외하고 2~6번째 row 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()  # 줄바꿈

# 셀 좌표에 대한 값을 가져올 필요가 있을 때
# row_range = ws[2:ws.max_row]  # 2번째 줄(row)부터 마지막 줄까지 가져오기
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)  # 튜플 형태로 가져옴
#         # print(xy, end=" ")
#         print(xy[0], end="")  # A와 같은 문자들
#         print(xy[1], end=" ")  # 1와 같은 숫자들
#     print()

# # 전체 행rows를 가져올 때
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[0].value)


# # 전체 열column을 가져올 때
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows():  # 전체 row에 대해 반복하면서 row를 가져오는 것
#     print(row[2].value)

# for column in ws.iter_cols():
#     print(column[0].value)


'''
 데이터를 행(row)단위로 가져오려면 iter_rows를 쓰면 되고,
 데이터를 열(column)단위로 가져오려면 iter_cols를 쓰면 된다.
'''
for row in ws.iter_rows(min_row=1, max_row=11, min_col=2, max_col=3):  # 가져오려 하는 셀의 범위 지정 가능
    print(row[0].value, row[1].value)  # iter_rows니까 행으로, 좌우좌우로 데이터를 가져옴
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)  # iter_cols니까 열로, 상하상하로 데이터를 가져옴

wb.save("Sample.xlsx")
