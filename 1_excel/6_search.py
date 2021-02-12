from openpyxl import load_workbook
from random import *
wb = load_workbook("Sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학 이런 식으로 데이터를 가져올 거
    # 여기서 영어점수가 80 이상인지 확인
    if int(row[1].value) >= 80:
        print(row[0].value, "번 학생은 영어 천재")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("Sample_modified.xlsx")
