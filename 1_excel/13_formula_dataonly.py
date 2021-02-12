from openpyxl import load_workbook
# 수식 그대로 가져오는 코드
# wb = load_workbook("Sample_formula.xlsx")

# 수식이 아닌, 계산된 결과 데이터를 그대로 가져오는 코드
wb = load_workbook("Sample_formula.xlsx", data_only=True)
ws = wb.active

# 이 엑셀에 있는 모든 데이터 가져오려는 코드
# 계산되지 않은 상태의 데이터는 None이라고 표시
for row in ws.values:
    for i in row:
        print(i)
