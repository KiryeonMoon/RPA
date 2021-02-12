from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])

scores = [
    (1, 10, 8, 5, 14, 26, 12),
    (2, 7, 3, 7, 15, 24, 18),
    (3, 9, 5, 8, 8, 12, 4),
    (4, 7, 8, 7, 17, 21, 18),
    (5, 7, 8, 7, 16, 25, 15),
    (6, 3, 5, 8, 8, 17, 0),
    (7, 4, 9, 10, 16, 27, 18),
    (8, 6, 6, 6, 15, 19, 17),
    (9, 10, 10, 9, 19, 30, 19),
    (10, 9, 8, 8, 20, 25, 20)
]

for i in scores:
    ws.append(i)

# 1. 퀴즈2 점수를 10으로 통일
for i in ws["D"]:
    if isinstance(i.value, int) == 1:
        i.value = 10

# 2. H열에 총점(Sum 이용), I열에 성적정보 추가
# 총점 90 이상 A ~ 총점 70 C, 나머지 D
ws["H1"] = "총점"
ws["I1"] = "성적"
for idx, row in enumerate(scores, start=2):
    grade = sum(row[1:]) - row[3] + 10
    ws.cell(row=idx, column=8).value = "=SUM(B{0}:G{1})".format(idx, idx)
    if grade >= 90:
        ws.cell(row=idx, column=9).value = "A"
    elif grade < 90 and grade >= 80:
        ws.cell(row=idx, column=9).value = "B"
    elif grade < 80 and grade >= 70:
        ws.cell(row=idx, column=9).value = "C"
    else:
        ws.cell(row=idx, column=9).value = "D"

# 3. 출석이 5 미만인 학생은 총점 상관없이 F
for idx, row in enumerate(scores, start=2):
    if ws.cell(row=idx, column=2).value < 5:
        ws.cell(row=idx, column=9).value = "F"

wb.save("quiz.xlsx")
