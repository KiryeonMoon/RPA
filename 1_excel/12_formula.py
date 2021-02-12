import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()  # 오늘 날짜정보
ws["A2"] = "=SUM(1,2,3)"  # 1,2,3이 더해진 6값이 나오는 수식
ws["A3"] = "=AVERAGE(1,2,3)"  # 1,2,3의 평균 2를 구하는 수식

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=sum(A4:A5)"  # A4셀부터 A5셀까지 합을 구하는 수식

wb.save("Sample_formula.xlsx")
