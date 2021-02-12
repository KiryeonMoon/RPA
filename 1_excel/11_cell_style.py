from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
wb = load_workbook("Sample.xlsx")
ws = wb.active

# 번호, 영어, 수학 각 셀 스타일 설정
a1 = ws["A1"]  # 번호
b1 = ws["B1"]  # 영어
c1 = ws["C1"]  # 수학

# A열의 너비 5로 설정
ws.column_dimensions["A"].width = 5
# 1행의 높이 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single")

# 테두리 그리기
thin_border = Border(left=Side(style="thin"), right=Side(
    style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 특정 셀에 대해 배경색 변경
# 90점 넘는 셀에 대해 초록색으로 적용
for row in ws.rows:
    for i in row:
        # 각 셀에 대해서 중앙정렬
        i.alignment = Alignment(horizontal="center", vertical="center")

        # A 번호열은 필요 없으니 제외
        if i.column == 1:
            continue  # 스킵하겠다고

        # value값이 정수형이고 90점보다 높으면
        if isinstance(i.value, int) and i.value >= 90:
            # 배경을 초록색으로 설정
            i.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            i.font = Font(color="FF0000")  # 폰트 색상도 변경

# 틀 고정
# B2 기준으로 틀 고정
ws.freeze_panes = "B2"

wb.save("Sample_style.xlsx")
