from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active
ws.title = "BirmanSheet"

# 셀 접근방법1
# A1 셀에 1이라는 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])  # A1셀 "정보" 출력
print(ws["A1"].value)  # A1셀 "값" 출력
print(ws["A10"].value)  # 값이 없을 땐 'None'을 출력


# 셀 접근방법2
# row = 1, 2, 3, ...
# column = A, B, C, ...
print(ws.cell(row=1, column=1).value)  # = wb["A1"]과 똑같은 의미
print(ws.cell(row=1, column=2).value)  # = wb["B1"]과 똑같은 의미

c = ws.cell(row=1, column=3, value=10)  # ws["C1"]=10과 동일한 코드
print(c.value)  # ws["C1"].value와 동일한 코드

# 반복문을 이용해서 랜덤숫자 채우기
index = 1
for x in range(1, 11):  # 10개 row
    for y in range(1, 11):  # 10개 column
        # ws.cell(row=x, column=y, value=randint())  # 0~100 사이의 숫자
        ws.cell(row=x, column=y, value=index)  # 0~100 사이의 숫자
        index += 1


wb.save("sample.xlsx")
