from openpyxl import load_workbook  # 파일 불러오는 역할
wb = load_workbook("sample.xlsx")  # sample.xlsx파일에서 wb 불러옴
ws = wb.active  # sheet 활성화

# # cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1, 11):
#         # end는 데이터를 1, 2, 3, .. 이렇게 찍고 싶어서 쓴 거
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):  # max_row는 최대 row 수를 의미!
    for y in range(1, ws.max_column+1):  # max_column은 최대 column수를 의미
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
