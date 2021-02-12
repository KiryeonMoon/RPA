from openpyxl import load_workbook
wb = load_workbook("Sample.xlsx")
ws = wb.active

# 번호 영어 수학 >> 번호 (국어) 영어 수학
# B1:C11 영역을, 줄 기준으로는 그대로, 열 기준으로 1칸 이동
ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어"  # B1 셀에 "국어" 입력

wb.save("Sample_move.xlsx")
