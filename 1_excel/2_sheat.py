from openpyxl import Workbook
wb = Workbook()

# wb.active랑 다르다.
# 현재 활성화된 시트 뒤에 새 시트를 기본이름으로 생성
ws = wb.create_sheet()
ws.title = "Mysheet"  # 시트이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # RGB 형태로 값을 넣음

# 주어진 이름으로 새 시트 바로 생성
ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2)  # 3번째 index에 sheet 생성

new_ws = wb["NewSheet"]  # Dictionary형태로 sheet에 접근 가능

print(wb.sheetnames)  # 모든 sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
